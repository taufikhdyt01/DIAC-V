import openpyxl
import re
from typing import Dict, Optional

class SimpleFormulaEvaluator:
    """
    Simple formula evaluator untuk mengatasi masalah formula di BDU View
    Hanya evaluasi formula yang umum digunakan
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        self.workbook_data = None
        
    def load_workbook(self) -> bool:
        """Load workbook untuk membaca formula dan data"""
        try:
            # Load workbook dengan formula (data_only=False)
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=False)
            # Load workbook dengan data calculated (data_only=True)
            self.workbook_data = openpyxl.load_workbook(self.excel_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading workbook: {str(e)}")
            return False
    
    def get_cell_value(self, sheet_name: str, cell_ref: str) -> str:
        """Get cell value dari workbook data"""
        try:
            if sheet_name in self.workbook_data.sheetnames:
                sheet = self.workbook_data[sheet_name]
                value = sheet[cell_ref].value
                return str(value) if value is not None else ""
        except Exception:
            return ""
        return ""
    
    def evaluate_concatenation(self, formula: str) -> str:
        """
        Evaluasi formula concatenation sederhana
        Contoh: ="td_Text " & 'Sheet'!B10 & ", " & 'Sheet'!B9
        """
        try:
            # Hapus = di awal
            formula = formula.strip()
            if formula.startswith('='):
                formula = formula[1:]
            
            # Split by & operator
            parts = re.split(r'\s*&\s*', formula)
            result_parts = []
            
            for part in parts:
                part = part.strip()
                
                # Handle string literals (dalam quotes)
                if (part.startswith('"') and part.endswith('"')):
                    literal_value = part[1:-1]  # Hapus quotes
                    result_parts.append(literal_value)
                
                # Handle cell references dengan sheet name
                elif "'" in part and '!' in part:
                    # Parse 'Sheet Name'!CellRef
                    quote_end = part.find("'", 1)
                    if quote_end > 0:
                        sheet_name = part[1:quote_end]
                        cell_ref = part[quote_end + 2:]  # Skip '!
                        cell_value = self.get_cell_value(sheet_name, cell_ref)
                        result_parts.append(cell_value)
                
                # Handle cell references tanpa quotes
                elif '!' in part:
                    sheet_name, cell_ref = part.split('!')
                    sheet_name = sheet_name.strip("'")
                    cell_value = self.get_cell_value(sheet_name, cell_ref)
                    result_parts.append(cell_value)
                
                else:
                    result_parts.append(part.strip('"\''))
            
            result = ''.join(result_parts)
            
            # Hapus prefix td_ atau tdi_
            if result.startswith("td_"):
                result = result[3:]
            elif result.startswith("tdi_"):
                result = result[4:]
            
            return result
            
        except Exception as e:
            print(f"Error evaluating concatenation: {str(e)}")
            return "[FORMULA ERROR]"
    
    def evaluate_if_simple(self, formula: str) -> str:
        """
        Evaluasi IF formula sederhana untuk payment terms
        """
        try:
            # Cari cell reference dalam IF
            cell_match = re.search(r"'([^']+)'!([A-Z]+\d+)", formula)
            if cell_match:
                sheet_name = cell_match.group(1)
                cell_ref = cell_match.group(2)
                cell_value = self.get_cell_value(sheet_name, cell_ref)
                
                # Mapping untuk payment terms
                if "14 Days" in cell_value or "14" in str(cell_value):
                    return "fourteen (14) days"
                elif "30 Days" in cell_value or "30" in str(cell_value):
                    return "thirty (30) days"
                elif "45 Days" in cell_value or "45" in str(cell_value):
                    return "forty-five (45) days"
                else:
                    return "[invalid payment term]"
            
            return "[IF ERROR]"
            
        except Exception as e:
            print(f"Error evaluating IF: {str(e)}")
            return "[IF ERROR]"
    
    def evaluate_simple_concat(self, formula: str, current_sheet: str) -> str:
        """
        Evaluasi concatenation sederhana seperti =J73&K73&L73
        """
        try:
            formula = formula.strip()
            if formula.startswith('='):
                formula = formula[1:]
            
            # Split by &
            parts = formula.split('&')
            result_parts = []
            
            for part in parts:
                part = part.strip()
                # Jika cell reference (contoh: J73)
                if re.match(r'^[A-Z]+\d+$', part):
                    cell_value = self.get_cell_value(current_sheet, part)
                    result_parts.append(cell_value)
                else:
                    result_parts.append(part)
            
            result = ''.join(result_parts)
            
            # Hapus prefix td_ atau tdi_
            if result.startswith("td_"):
                result = result[3:]
            elif result.startswith("tdi_"):
                result = result[4:]
                
            return result
            
        except Exception as e:
            print(f"Error evaluating simple concat: {str(e)}")
            return "[CONCAT ERROR]"
    
    def get_evaluated_value(self, sheet_name: str, cell_ref: str) -> str:
        """
        Main method untuk mendapatkan nilai yang sudah dievaluasi
        """
        try:
            # Coba ambil formula dulu
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                cell = sheet[cell_ref]
                
                if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Tentukan jenis formula dan evaluasi
                    if '&' in formula and 'IF(' not in formula:
                        if '!' in formula:
                            return self.evaluate_concatenation(formula)
                        else:
                            return self.evaluate_simple_concat(formula, sheet_name)
                    elif formula.strip().startswith('=IF('):
                        return self.evaluate_if_simple(formula)
                    else:
                        # Untuk formula complex lainnya, coba ambil calculated value
                        calculated_value = self.get_cell_value(sheet_name, cell_ref)
                        return calculated_value if calculated_value else "[COMPLEX FORMULA]"
                else:
                    # Bukan formula, ambil nilai biasa
                    value = self.get_cell_value(sheet_name, cell_ref)
                    # Hapus prefix jika ada
                    if isinstance(value, str):
                        if value.startswith("td_"):
                            return value[3:]
                        elif value.startswith("tdi_"):
                            return value[4:]
                    return value
        
        except Exception as e:
            print(f"Error getting evaluated value for {sheet_name}!{cell_ref}: {str(e)}")
            return "[ERROR]"
    
    def close(self):
        """Tutup workbooks"""
        if self.workbook:
            self.workbook.close()
        if self.workbook_data:
            self.workbook_data.close()

# Configuration untuk cell-cell yang mengandung formula
FORMULA_CELLS = {
    # Format: "sheet_name.cell_ref": "display_name"
    "DIP_Project Information.B85": "Cost and Freight Information",
    "DIP_Project Information.B86": "Payment Terms Information",
    # Tambahkan cell formula lainnya di sini jika diperlukan
}

def evaluate_formulas_background(evaluator):
    """
    Evaluasi semua formula di background tanpa menampilkan UI
    Hanya untuk memastikan formula ter-process dengan benar
    """
    try:
        # Process semua formula cells yang dikonfigurasi
        processed_count = 0
        for full_cell_ref, display_name in FORMULA_CELLS.items():
            sheet_name, cell_ref = full_cell_ref.split(".", 1)
            
            try:
                # Evaluasi formula
                result = evaluator.get_evaluated_value(sheet_name, cell_ref)
                processed_count += 1
            except Exception as e:
                print(f"❌ Error processing {display_name}: {str(e)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error in background formula processing: {str(e)}")
        return False

def create_formula_widget(parent, evaluator, sheet_name, cell_ref, display_name):
    """Create widget untuk menampilkan hasil formula"""
    from PyQt5.QtWidgets import QFrame, QVBoxLayout, QLabel
    from PyQt5.QtGui import QFont
    from PyQt5.QtCore import Qt
    
    frame = QFrame()
    frame.setStyleSheet("""
        QFrame {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            padding: 10px;
            margin: 5px;
        }
    """)
    
    layout = QVBoxLayout(frame)
    layout.setContentsMargins(10, 8, 10, 8)
    
    # Label nama field
    label = QLabel(display_name)
    label.setFont(QFont("Segoe UI", 11, QFont.Bold))
    label.setStyleSheet("color: #495057; margin-bottom: 5px;")
    
    # Label konten
    content_label = QLabel()
    content_label.setFont(QFont("Segoe UI", 10))
    content_label.setStyleSheet("""
        color: #212529; 
        background-color: white; 
        padding: 8px; 
        border-radius: 3px;
        border: 1px solid #e9ecef;
    """)
    content_label.setWordWrap(True)
    content_label.setMinimumHeight(35)
    
    # Evaluasi dan set nilai
    try:
        evaluated_value = evaluator.get_evaluated_value(sheet_name, cell_ref)
        content_label.setText(evaluated_value)
        
        # Update style jika error
        if evaluated_value.startswith("[") and evaluated_value.endswith("]"):
            content_label.setStyleSheet("""
                color: #dc3545; 
                background-color: #fff5f5; 
                padding: 8px; 
                border-radius: 3px;
                border: 1px solid #f5c6cb;
            """)
    except Exception as e:
        content_label.setText(f"[Error: {str(e)}]")
        content_label.setStyleSheet("""
            color: #dc3545; 
            background-color: #fff5f5; 
            padding: 8px; 
            border-radius: 3px;
            border: 1px solid #f5c6cb;
        """)
    
    layout.addWidget(label)
    layout.addWidget(content_label)
    
    # Store references untuk refresh
    frame.content_label = content_label
    frame.sheet_name = sheet_name
    frame.cell_ref = cell_ref
    frame.evaluator = evaluator
    
    return frame