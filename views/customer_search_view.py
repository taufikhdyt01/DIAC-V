import os
import sys
import re
import shutil
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QPushButton, QFrame, QGridLayout, QSpacerItem,
                             QSizePolicy, QScrollArea, QApplication, QMenu, QAction,
                             QLineEdit, QListWidget, QListWidgetItem, QMessageBox, QProgressDialog)
from PyQt5.QtGui import QPixmap, QIcon, QFont, QColor, QPalette, QCursor
from PyQt5.QtCore import Qt, QSize, pyqtSignal, QPoint, QTimer

# Import local modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import APP_NAME, APP_LOGO, DEPARTMENTS, PRIMARY_COLOR, SECONDARY_COLOR, BG_COLOR
from views.loading_screen import LoadingScreen, QuickLoadingDialog, show_loading_dialog

import pandas as pd

def clean_folder_name(name):
    """
    Create a valid Windows folder name by:
    1. Removing trailing whitespace and newlines
    2. Replacing invalid characters with underscores
    3. Ensuring name isn't too long for Windows paths
    """
    # Trim whitespace and newlines
    name = name.strip()
    
    # Replace invalid Windows filename characters: \ / : * ? " < > |
    name = re.sub(r'[\\/:*?"<>|\t\n\r]', '_', name)
    
    # Replace multiple consecutive underscores with a single one
    name = re.sub(r'_+', '_', name)
    
    # Limit the length to avoid path too long errors (Windows MAX_PATH is 260)
    # Use a reasonable limit like 100 chars for the folder name
    if len(name) > 100:
        name = name[:97] + '...'
    
    return name

class CustomerSearchView(QMainWindow):
    """View untuk pencarian customer sebelum akses ke BDU View"""
    
    # Signals
    back_to_dashboard = pyqtSignal()
    open_bdu_view = pyqtSignal(str)  # Pass customer_name as parameter
    
    def __init__(self, auth_manager):
        super().__init__()
        self.auth_manager = auth_manager
        self.current_user = auth_manager.get_current_user()
        self.customers_data = []
        self.filtered_customers = []
        self.selected_customer = None
        
        # Path to customer database
        self.db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                 "data", "database_customer.xlsx")
                                 
        # Path to SET_BDU template
        self.template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                     "data", "SET_BDU.xlsx")
                                     
        # Base path for customer folders
        self.customers_base_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                          "data", "customers")
        
        # Create customers directory if it doesn't exist
        if not os.path.exists(self.customers_base_path):
            os.makedirs(self.customers_base_path)
        
        self.initUI()
        self.load_customer_data()
    
    def initUI(self):
        """Initialize UI elements"""
        # Set window properties
        self.setWindowTitle(f"Customer Search - {APP_NAME}")
        self.setMinimumSize(1000, 700)
        
        # Set central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Header
        self.setup_header()
        main_layout.addWidget(self.header_widget)
        
        # Content area
        content_widget = QWidget()
        content_widget.setStyleSheet(f"background-color: {BG_COLOR};")
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # Add a white panel for the search area
        search_panel = QWidget()
        search_panel.setStyleSheet("background-color: white;")
        search_layout = QVBoxLayout(search_panel)
        search_layout.setContentsMargins(40, 40, 40, 40)
        search_layout.setSpacing(20)
        
        # Title and back button section
        title_layout = QHBoxLayout()
        
        # Back button
        back_btn = QPushButton()
        back_btn.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_ArrowBack))
        back_btn.setFixedSize(36, 36)
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {PRIMARY_COLOR};
                border-radius: 18px;
                color: white;
            }}
            QPushButton:hover {{
                background-color: #1f2c39;
            }}
        """)
        back_btn.clicked.connect(self.go_back_to_dashboard)
        
        # BDU title with icon
        bdu_dept = next((dept for dept in DEPARTMENTS if dept["id"] == "BDU"), None)
        if bdu_dept:
            dept_icon = QLabel(bdu_dept["emoji"])
            dept_icon.setFont(QFont("Segoe UI", 22))
            dept_icon.setStyleSheet(f"color: {bdu_dept['color']};")
            title_layout.addWidget(dept_icon)
        
        page_title = QLabel("BDU Customer Search")
        page_title.setFont(QFont("Segoe UI", 22, QFont.Bold))
        page_title.setStyleSheet(f"color: {PRIMARY_COLOR};")
        
        title_layout.addWidget(back_btn)
        title_layout.addSpacing(10)
        title_layout.addWidget(page_title)
        title_layout.addStretch()
        
        search_layout.addLayout(title_layout)
        
        # Subtitle
        subtitle = QLabel("Search for existing customer or register a new one")
        subtitle.setFont(QFont("Segoe UI", 12))
        subtitle.setStyleSheet("color: #666;")
        search_layout.addWidget(subtitle)
        
        # Search box
        search_box_layout = QVBoxLayout()
        search_box_layout.setSpacing(10)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Start typing customer name...")
        self.search_input.setFont(QFont("Segoe UI", 14))
        self.search_input.setMinimumHeight(50)
        self.search_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #ccc;
                border-radius: 5px;
                padding: 10px 15px;
                background-color: #f9f9f9;
            }
            QLineEdit:focus {
                border: 1px solid #3498db;
                background-color: white;
            }
        """)
        self.search_input.textChanged.connect(self.filter_customers)
        
        search_box_layout.addWidget(self.search_input)
        
        # Results list
        self.results_list = QListWidget()
        self.results_list.setMinimumHeight(300)  # Make sure list is visible
        self.results_list.setFont(QFont("Segoe UI", 12))
        self.results_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #ccc;
                border-radius: 5px;
                padding: 5px;
                background-color: white;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #eee;
            }
            QListWidget::item:selected {
                background-color: #e0f0ff;
                color: #333;
            }
            QListWidget::item:hover {
                background-color: #f0f0f0;
            }
        """)
        self.results_list.itemClicked.connect(self.on_customer_selected)
        self.results_list.itemDoubleClicked.connect(self.on_customer_double_clicked)
        
        search_box_layout.addWidget(self.results_list)
        
        # No results message
        self.no_results_label = QLabel("No matching customers found")
        self.no_results_label.setFont(QFont("Segoe UI", 12))
        self.no_results_label.setStyleSheet("color: #666; margin-top: 10px;")
        self.no_results_label.setAlignment(Qt.AlignCenter)
        self.no_results_label.setVisible(False)
        
        search_box_layout.addWidget(self.no_results_label)
        
        search_layout.addLayout(search_box_layout)
        
        # Buttons section
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        # Register New Customer button
        self.register_btn = QPushButton("Register New Customer")
        self.register_btn.setFont(QFont("Segoe UI", 12))
        self.register_btn.setMinimumHeight(40)
        self.register_btn.setCursor(Qt.PointingHandCursor)
        self.register_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {SECONDARY_COLOR};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
            }}
            QPushButton:hover {{
                background-color: #2980b9;
            }}
        """)
        self.register_btn.clicked.connect(self.register_new_customer)
        
        # Continue button (disabled initially)
        self.continue_btn = QPushButton("Continue to BDU Form")
        self.continue_btn.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.continue_btn.setMinimumHeight(40)
        self.continue_btn.setCursor(Qt.PointingHandCursor)
        self.continue_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {PRIMARY_COLOR};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
            }}
            QPushButton:hover {{
                background-color: #1f2c39;
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
                color: #999999;
            }}
        """)
        self.continue_btn.setEnabled(False)
        self.continue_btn.clicked.connect(self.continue_to_bdu)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.register_btn)
        buttons_layout.addWidget(self.continue_btn)
        
        search_layout.addLayout(buttons_layout)
        
        # Add search panel to content
        content_layout.addWidget(search_panel)
        
        # Add content to main layout
        main_layout.addWidget(content_widget)
        
        # Status bar
        self.statusBar().showMessage(f"BDU Customer Search | User: {self.current_user['username']}")
        self.statusBar().setStyleSheet("background-color: #f0f0f0; color: #555;")
        
        # Setup a timer for delayed search (for live search)
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.perform_search)
    
    def setup_header(self):
        """Setup header widget dengan logo, judul, dan menu"""
        self.header_widget = QWidget()
        self.header_widget.setFixedHeight(60)
        self.header_widget.setStyleSheet(f"background-color: {PRIMARY_COLOR};")
        
        header_layout = QHBoxLayout(self.header_widget)
        header_layout.setContentsMargins(15, 0, 15, 0)
        
        # BDU title with icon
        bdu_dept = next((dept for dept in DEPARTMENTS if dept["id"] == "BDU"), None)
        if bdu_dept:
            dept_icon = QLabel(bdu_dept["emoji"])
            dept_icon.setFont(QFont("Segoe UI", 18))
            dept_icon.setStyleSheet("color: white;")
            header_layout.addWidget(dept_icon)
        
        # App title
        title_label = QLabel(f"{APP_NAME} - BDU Group")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setStyleSheet("color: white;")
        
        # Spacer
        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        
        # User button
        user_menu_btn = QPushButton("👤")
        user_menu_btn.setFont(QFont("Segoe UI", 14))
        user_menu_btn.setStyleSheet("""
            QPushButton {
                border: none;
                background-color: transparent;
                color: white;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.2);
                border-radius: 5px;
            }
        """)
        user_menu_btn.setFixedSize(36, 36)
        user_menu_btn.setCursor(Qt.PointingHandCursor)
        user_menu_btn.clicked.connect(self.show_user_menu)
        
        # Add all elements to header layout
        header_layout.addWidget(title_label)
        header_layout.addItem(spacer)
        header_layout.addWidget(user_menu_btn)
    
    def show_user_menu(self):
        """Tampilkan menu pengguna"""
        sender = self.sender()
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 5px;
            }
            QMenu::item {
                padding: 5px 15px;
                border-radius: 3px;
            }
            QMenu::item:selected {
                background-color: #f0f0f0;
            }
        """)
        
        # Add user info at the top (non-clickable)
        user_info = QAction(f"{self.current_user['name']} ({self.current_user['department']})", self)
        user_info.setEnabled(False)
        menu.addAction(user_info)
        
        menu.addSeparator()
        
        # Add menu actions
        dashboard_action = QAction("Back to Dashboard", self)
        settings_action = QAction("Settings", self)
        
        menu.addAction(dashboard_action)
        menu.addAction(settings_action)
        
        # Connect actions
        dashboard_action.triggered.connect(self.go_back_to_dashboard)
        
        # Show menu at button position
        menu.exec_(sender.mapToGlobal(QPoint(0, sender.height())))
    
    def go_back_to_dashboard(self):
        """Go back to dashboard"""
        self.back_to_dashboard.emit()
        self.close()
    
    def load_customer_data(self):
        """Load customer data from database_customer.xlsx"""
        try:
            if os.path.exists(self.db_path):
                # Load the Excel file
                df = pd.read_excel(self.db_path)
                
                # Get list of customer names from the 'Company Name' column
                if 'Company Name' in df.columns:
                    self.customers_data = df['Company Name'].dropna().unique().tolist()
                    self.customers_data.sort()  # Sort alphabetically
                    
                    # Show all customers initially
                    self.filtered_customers = self.customers_data.copy()
                    self.update_results_list()
                    
                    # Update status message
                    self.statusBar().showMessage(f"Loaded {len(self.customers_data)} customers from database")
                else:
                    self.statusBar().showMessage("Error: 'Company Name' column not found in customer database")
                    self.no_results_label.setText("Error: 'Company Name' column not found in database")
                    self.no_results_label.setVisible(True)
            else:
                self.statusBar().showMessage(f"Customer database not found: {self.db_path}")
                self.no_results_label.setText("Customer database file not found")
                self.no_results_label.setVisible(True)
        except Exception as e:
            self.statusBar().showMessage(f"Error loading customer data: {str(e)}")
            self.no_results_label.setText(f"Error loading customer database: {str(e)}")
            self.no_results_label.setVisible(True)
            print(f"Error loading customer data: {str(e)}")
    
    def filter_customers(self):
        """Filter customers based on search input (with delay for performance)"""
        # Reset timer and start it again (300ms delay)
        self.search_timer.stop()
        self.search_timer.start(300)
    
    def perform_search(self):
        """Perform the actual search after delay"""
        search_text = self.search_input.text().strip().lower()
        
        if not search_text:
            # Show all customers if search is empty
            self.filtered_customers = self.customers_data.copy()
        else:
            # Filter customers that start with or contain the search text
            self.filtered_customers = [
                customer for customer in self.customers_data 
                if search_text in customer.lower()
            ]
        
        # Update the results list
        self.update_results_list()
    
    def update_results_list(self):
        """Update the results list with filtered customers"""
        self.results_list.clear()
        
        if self.filtered_customers:
            # Add filtered customers to list
            for customer in self.filtered_customers:
                item = QListWidgetItem(customer)
                
                # Check if customer has existing BDU file and add icon indicator
                if self.check_customer_bdu_file_exists(customer):
                    item.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_FileIcon))
                    item.setToolTip(f"Customer has existing BDU file")
                
                self.results_list.addItem(item)
            
            # Hide no results message
            self.no_results_label.setVisible(False)
        else:
            # Show no results message
            if self.search_input.text().strip():
                self.no_results_label.setText(f"No customers found matching '{self.search_input.text()}'")
            else:
                self.no_results_label.setText("No customers in database")
            
            self.no_results_label.setVisible(True)
    
    def on_customer_selected(self, item):
        """Handle customer selection from the list"""
        self.selected_customer = item.text()
        self.continue_btn.setEnabled(True)
        
        # Check if customer file exists and update status bar
        file_exists = self.check_customer_bdu_file_exists(self.selected_customer)
        status_message = f"Selected customer: {self.selected_customer}"
        
        if file_exists:
            status_message += " (existing BDU file will be opened)"
        else:
            status_message += " (new BDU file will be created)"
            
        self.statusBar().showMessage(status_message)
    
    def on_customer_double_clicked(self, item):
        """Handle customer double click (select and continue)"""
        self.selected_customer = item.text()
        self.continue_to_bdu()
    
    def register_new_customer(self):
        """Register a new customer"""
        # Get customer name from search input if not empty
        new_customer = self.search_input.text().strip()
        
        if not new_customer:
            QMessageBox.warning(self, "Input Required", "Please enter a customer name to register")
            return
        
        # Check if customer already exists
        if new_customer in self.customers_data:
            QMessageBox.information(
                self, 
                "Customer Exists", 
                f"Customer '{new_customer}' already exists in the database.\n\nPlease select it from the list."
            )
            
            # Find and select that customer in the list
            for i in range(self.results_list.count()):
                if self.results_list.item(i).text() == new_customer:
                    self.results_list.setCurrentRow(i)
                    self.selected_customer = new_customer
                    self.continue_btn.setEnabled(True)
                    break
        else:
            # Confirm registration
            reply = QMessageBox.question(
                self,
                "Register New Customer",
                f"Register '{new_customer}' as a new customer?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Show loading screen for registration process
                def register_customer_process(progress_callback=None):
                    try:
                        if progress_callback:
                            progress_callback(20, f"Creating customer folder for {new_customer}...")
                        
                        # Create customer folder in advance
                        self.get_customer_folder_path(new_customer)
                        
                        if progress_callback:
                            progress_callback(60, "Adding customer to database...")
                        
                        # Add new customer to the database if it exists
                        self.add_customer_to_database(new_customer)
                        
                        if progress_callback:
                            progress_callback(80, "Updating customer list...")
                        
                        # Add new customer to temporary list
                        self.customers_data.append(new_customer)
                        self.filtered_customers.append(new_customer)
                        
                        if progress_callback:
                            progress_callback(100, "Customer registration completed!")
                        
                        return True
                    except Exception as e:
                        if progress_callback:
                            progress_callback(100, f"Error: {str(e)}")
                        return False
                
                # Show loading screen
                loading_screen = LoadingScreen(
                    parent=self,
                    title="Registering Customer",
                    message=f"Setting up new customer: {new_customer}"
                )
                loading_screen.show()
                loading_screen.start_loading(register_customer_process)
                
                # Connect completion handler
                def on_registration_complete(success, message):
                    if success:
                        # Set as selected customer and enable continue
                        self.selected_customer = new_customer
                        self.continue_btn.setEnabled(True)
                        
                        self.update_results_list()
                        
                        # Find and select the new customer in the list
                        for i in range(self.results_list.count()):
                            if self.results_list.item(i).text() == new_customer:
                                self.results_list.setCurrentRow(i)
                                break
                        
                        self.statusBar().showMessage(f"New customer '{new_customer}' registered and added to database")
                        
                        # Show success message
                        QTimer.singleShot(500, lambda: QMessageBox.information(
                            self, "Registration Complete", 
                            f"Customer '{new_customer}' has been successfully registered!"
                        ))
                    else:
                        QMessageBox.critical(self, "Registration Failed", f"Failed to register customer: {message}")
                
                loading_screen.worker.task_completed.connect(on_registration_complete)
    
    def add_customer_to_database(self, customer_name):
        """Add new customer to database_customer.xlsx"""
        try:
            if os.path.exists(self.db_path):
                # Load existing database
                df = pd.read_excel(self.db_path)
                
                # Check if 'Company Name' column exists
                if 'Company Name' in df.columns:
                    # Create a new row for the customer
                    new_row = pd.DataFrame({
                        'Company Name': [customer_name]
                    })
                    
                    # Fill other columns with empty values if they exist
                    for col in df.columns:
                        if col != 'Company Name':
                            new_row[col] = ""
                    
                    # Append to the dataframe
                    df = pd.concat([df, new_row], ignore_index=True)
                    
                    # Save back to Excel
                    df.to_excel(self.db_path, index=False)
                    print(f"Added customer '{customer_name}' to database")
        except Exception as e:
            print(f"Error adding customer to database: {str(e)}")
            # Don't show error to user, just log it - customer can still be used

    def get_customer_folder_path(self, customer_name):
        """Get the path to a customer's folder, creating it if it doesn't exist"""
        # Validate input - make sure customer_name is a string
        if not isinstance(customer_name, str):
            raise TypeError(f"Customer name must be a string, got {type(customer_name)}")
            
        # Create a valid folder name from the customer name
        valid_folder_name = clean_folder_name(customer_name)
        folder_path = os.path.join(self.customers_base_path, valid_folder_name)
        
        return folder_path
    
    def get_customer_bdu_file_path(self, customer_name):
        """Get the path to a customer's SET_BDU.xlsx file"""
        # Validate input - make sure customer_name is a string
        if not isinstance(customer_name, str):
            raise TypeError(f"Customer name must be a string, got {type(customer_name)}")
            
        folder_path = self.get_customer_folder_path(customer_name)
        return os.path.join(folder_path, "SET_BDU.xlsx")
    
    def check_customer_bdu_file_exists(self, customer_name):
        """Check if the customer's SET_BDU.xlsx file already exists"""
        # Validate input - make sure customer_name is a string
        if not isinstance(customer_name, str):
            raise TypeError(f"Customer name must be a string, got {type(customer_name)}")
            
        file_path = self.get_customer_bdu_file_path(customer_name)
        return os.path.exists(file_path)
    
    def create_customer_bdu_file(self, customer_name, progress_callback=None):
        """Create a copy of SET_BDU.xlsx for the customer if it doesn't exist"""
        # Get paths
        customer_file_path = self.get_customer_bdu_file_path(customer_name)
        
        # Check if template exists
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        # Check if file already exists
        if os.path.exists(customer_file_path):
            print(f"Customer BDU file already exists: {customer_file_path}")
            return customer_file_path
        
        try:
            if progress_callback:
                progress_callback(10, "Loading customer data from database...")
            
            # First, get all customer data from the database
            customer_data = self.get_customer_data_from_database(customer_name)
            if not customer_data:
                print(f"Warning: Could not find detailed data for customer '{customer_name}' in database")
            
            if progress_callback:
                progress_callback(30, "Copying template file...")
            
            # Copy the template file to the customer folder
            shutil.copy2(self.template_path, customer_file_path)
            
            if progress_callback:
                progress_callback(50, "Updating customer information...")
            
            # Update customer information in the Excel file
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(customer_file_path)
                
                # Specifically target the DIP_Customer Information sheet
                sheet_name = 'DIP_Customer Information'
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    if progress_callback:
                        progress_callback(70, "Mapping customer data to Excel cells...")
                    
                    # Update customer data in specific cells
                    data_mapping = {
                        'B4': customer_name,  # Company Name
                        'B8': customer_data.get('Country', ''),
                        'B9': customer_data.get('Province', ''),
                        'B10': customer_data.get('City', ''),
                        'B11': customer_data.get('Site Address', ''),
                        'B12': customer_data.get('Correspondence Address (HO)', ''),
                        'B13': customer_data.get('Postal Code (HO)', ''),
                        'B16': customer_data.get('Name', ''),
                        'C16': customer_data.get('Phone No./Email', '')
                    }
                    
                    # Apply all mappings
                    for cell, value in data_mapping.items():
                        if value:  # Only update if we have a value
                            sheet[cell] = value
                            print(f"Updated {sheet_name}.{cell} with: {value}")
                
                if progress_callback:
                    progress_callback(85, "Updating additional sheets...")
                
                # Also try to update customer name in various other sheets where it might be expected
                sheets_to_check = ['DATA_GENERAL', 'DATA_CUSTOMER', 'DIP_General Information']
                customer_fields = ['customer', 'client', 'company']
                
                for check_sheet_name in sheets_to_check:
                    if check_sheet_name in workbook.sheetnames:
                        sheet = workbook[check_sheet_name]
                        
                        # Search for customer name fields in the first 30 rows
                        for row in range(1, min(30, sheet.max_row + 1)):
                            for col in range(1, min(3, sheet.max_column + 1)):
                                cell_value = sheet.cell(row=row, column=col).value
                                
                                # Check if this is a cell that might need customer name
                                if isinstance(cell_value, str) and any(field in cell_value.lower() for field in customer_fields):
                                    # Update the cell to the right (B column if A contains customer)
                                    next_col = col + 1
                                    if next_col <= sheet.max_column:
                                        sheet.cell(row=row, column=next_col).value = customer_name
                                        print(f"Updated customer name in sheet {check_sheet_name} at cell {chr(64+next_col)}{row}")
                
                if progress_callback:
                    progress_callback(95, "Saving customer file...")
                
                # Save the workbook
                workbook.save(customer_file_path)
                
                if progress_callback:
                    progress_callback(100, "Customer file created successfully!")
                
            except Exception as e:
                print(f"Error updating customer data in Excel: {str(e)}")
                import traceback
                traceback.print_exc()
                # Continue anyway - the file was copied
            
            return customer_file_path
            
        except Exception as e:
            raise Exception(f"Error creating customer BDU file: {str(e)}")

    def get_customer_data_from_database(self, customer_name):
        """Retrieve all customer data from database_customer.xlsx"""
        try:
            if not os.path.exists(self.db_path):
                print(f"Customer database not found: {self.db_path}")
                return {}
                
            # Load the database
            df = pd.read_excel(self.db_path)
            
            # Find the customer row
            customer_rows = df[df['Company Name'] == customer_name]
            if customer_rows.empty:
                print(f"Customer '{customer_name}' not found in database")
                return {}
                
            # Get the first matching row (in case of duplicates)
            customer_row = customer_rows.iloc[0]
            
            # Convert the row to a dictionary
            customer_data = customer_row.to_dict()
            
            # Clean up the data - handle NaN values
            for key, value in customer_data.items():
                if pd.isna(value):
                    customer_data[key] = ""
                elif not isinstance(value, str):
                    # Convert numeric values to strings
                    customer_data[key] = str(value)
                    
            return customer_data
            
        except Exception as e:
            print(f"Error getting customer data from database: {str(e)}")
            import traceback
            traceback.print_exc()
            return {}
    
    def continue_to_bdu(self):
        """Continue to BDU View with selected customer"""
        if not self.selected_customer:
            QMessageBox.warning(self, "Selection Required", "Please select a customer to continue")
            return
        
        def prepare_bdu_workspace(progress_callback=None):
            try:
                if progress_callback:
                    progress_callback(10, "Checking template file...")
                
                # Check if template file exists
                if not os.path.exists(self.template_path):
                    raise Exception(f"Template file not found: {self.template_path}\n\nPlease ensure SET_BDU.xlsx exists in the data folder.")
                
                if progress_callback:
                    progress_callback(30, f"Checking customer file for {self.selected_customer}...")
                
                # Check if customer BDU file exists, create if it doesn't
                if not self.check_customer_bdu_file_exists(self.selected_customer):
                    if progress_callback:
                        progress_callback(40, f"Creating new BDU file for {self.selected_customer}...")
                    
                    # Create the file with progress updates
                    self.create_customer_bdu_file(self.selected_customer, progress_callback)
                else:
                    if progress_callback:
                        progress_callback(80, f"Found existing BDU file for {self.selected_customer}")
                
                if progress_callback:
                    progress_callback(100, "BDU workspace ready!")
                
                return True
            except Exception as e:
                if progress_callback:
                    progress_callback(100, f"Error: {str(e)}")
                return False
        
        # Show loading screen
        loading_screen = LoadingScreen(
            parent=self,
            title="Preparing BDU Workspace",
            message=f"Setting up BDU workspace for {self.selected_customer}..."
        )
        loading_screen.show()
        loading_screen.start_loading(prepare_bdu_workspace)
        
        # Connect completion handler
        def on_preparation_complete(success, message):
            if success:
                # Signal to open BDU view with selected customer
                QTimer.singleShot(300, lambda: self.open_bdu_view.emit(self.selected_customer))
                self.close()
            else:
                error_msg = message if "Error:" in message else f"An error occurred while preparing the customer file:\n\n{message}"
                QMessageBox.critical(self, "Error", error_msg)
        
        loading_screen.worker.task_completed.connect(on_preparation_complete)